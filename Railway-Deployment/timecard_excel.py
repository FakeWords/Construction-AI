"""
Excel Exporter for Time Card Data
Creates Excel file with separate FTE and Contractor sheets
"""

from typing import List, Dict
import io


def create_timecard_excel(fte_entries: List[Dict], contractor_entries: List[Dict]) -> bytes:
    """
    Create Excel file with two sheets: FTE and Contractor
    
    Columns:
    - Vendor/Tesla Employee ID
    - Full Name
    - Scope ID
    - PO #
    - PO Line#
    - Time In
    - Time Out
    - Lunch
    - Day/Night Shift
    - Hours Worked
    - Comments
    """
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("openpyxl not installed - required for Excel export")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create FTE sheet
    ws_fte = wb.create_sheet('FTE')
    setup_timecard_sheet(ws_fte, fte_entries, 'Tesla Employee ID')
    
    # Create Contractor sheet
    ws_contractor = wb.create_sheet('Contractor')
    setup_timecard_sheet(ws_contractor, contractor_entries, 'Vendor Employee ID')
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


def setup_timecard_sheet(ws, entries: List[Dict], id_column_name: str):
    """Setup worksheet with headers and data"""
    
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Headers
    headers = [
        id_column_name,
        'Full Name',
        'Scope ID',
        'PO #',
        'PO Line#',
        'Time In',
        'Time Out',
        'Lunch',
        'Day/Night Shift',
        'Hours Worked',
        'Comments'
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    column_widths = {
        'A': 18,  # ID
        'B': 25,  # Full Name
        'C': 12,  # Scope ID
        'D': 18,  # PO #
        'E': 12,  # PO Line#
        'F': 12,  # Time In
        'G': 12,  # Time Out
        'H': 10,  # Lunch
        'I': 15,  # Shift
        'J': 14,  # Hours Worked
        'K': 30   # Comments
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Write data rows
    for row_idx, entry in enumerate(entries, start=2):
        # Data alignment
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Employee ID
        cell = ws.cell(row=row_idx, column=1, value=entry['employee_id'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # Full Name
        cell = ws.cell(row=row_idx, column=2, value=entry['full_name'])
        cell.alignment = left_align
        cell.border = thin_border
        
        # Scope ID
        cell = ws.cell(row=row_idx, column=3, value=entry['scope_id'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # PO #
        cell = ws.cell(row=row_idx, column=4, value=entry['po_number'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # PO Line#
        cell = ws.cell(row=row_idx, column=5, value=entry['po_line'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # Time In
        cell = ws.cell(row=row_idx, column=6, value=entry['time_in'])
        cell.alignment = center_align
        cell.border = thin_border
        if not entry['time_in']:
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # Time Out
        cell = ws.cell(row=row_idx, column=7, value=entry['time_out'])
        cell.alignment = center_align
        cell.border = thin_border
        if not entry['time_out']:
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        # Lunch
        cell = ws.cell(row=row_idx, column=8, value=entry['lunch'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # Shift
        cell = ws.cell(row=row_idx, column=9, value=entry['shift'])
        cell.alignment = center_align
        cell.border = thin_border
        
        # Hours Worked
        hours_val = entry['hours_worked'] if entry['hours_worked'] else '0.00'
        cell = ws.cell(row=row_idx, column=10, value=hours_val)
        cell.alignment = center_align
        cell.border = thin_border
        cell.number_format = '0.00'
        
        # Highlight if hours are unusual
        if entry['hours_worked']:
            hours_float = float(entry['hours_worked'])
            if hours_float > 12 or hours_float < 1:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Comments
        cell = ws.cell(row=row_idx, column=11, value=entry['comments'])
        cell.alignment = left_align
        cell.border = thin_border
        if entry['comments']:
            cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Add summary row at bottom
    if entries:
        summary_row = len(entries) + 3
        
        ws.cell(row=summary_row, column=1, value='TOTAL HOURS:').font = Font(bold=True)
        
        # Sum hours
        total_hours = sum(float(e['hours_worked']) if e['hours_worked'] else 0 for e in entries)
        cell = ws.cell(row=summary_row, column=10, value=total_hours)
        cell.font = Font(bold=True)
        cell.number_format = '0.00'
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
