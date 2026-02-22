"""
Excel Material Takeoff Export
Generates professional spreadsheets from drawing analysis
Format: Same as what estimators create manually
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List
import re


class MaterialTakeoffExporter:
    """
    Creates professional Excel takeoff spreadsheets
    Formatted like real estimators use
    """
    
    def __init__(self, project_name: str = "Construction Project"):
        self.project_name = project_name
        self.wb = Workbook()
        
        # Professional color scheme
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.subheader_font = Font(bold=True, color="FFFFFF", size=11)
        self.total_font = Font(bold=True, size=11)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_electrical_takeoff(self, analysis_result: str, drawing_name: str) -> str:
        """
        Create electrical material takeoff from analysis
        
        Returns: Path to Excel file
        """
        # Parse the analysis to extract quantities
        equipment, wire, conduit = self._parse_electrical_analysis(analysis_result)
        
        # Create summary sheet
        ws_summary = self.wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, drawing_name)
        
        # Create equipment sheet
        ws_equipment = self.wb.create_sheet("Equipment")
        self._create_equipment_sheet(ws_equipment, equipment)
        
        # Create wire sheet
        ws_wire = self.wb.create_sheet("Wire & Cable")
        self._create_wire_sheet(ws_wire, wire)
        
        # Create conduit sheet
        ws_conduit = self.wb.create_sheet("Conduit & Fittings")
        self._create_conduit_sheet(ws_conduit, conduit)
        
        # Save
        filename = f"Takeoff_{drawing_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.wb.save(filename)
        
        return filename
    
    def _create_summary_sheet(self, ws, drawing_name):
        """Create professional summary sheet"""
        # Title
        ws['A1'] = f"{self.project_name} - Material Takeoff"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Project info
        ws['A3'] = "Drawing:"
        ws['B3'] = drawing_name
        ws['A4'] = "Date:"
        ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws['A5'] = "Prepared by:"
        ws['B5'] = "Fieldwise AI - Automated Takeoff"
        
        # Summary header
        ws['A7'] = "SUMMARY"
        ws['A7'].font = self.header_font
        ws['A7'].fill = self.header_fill
        ws.merge_cells('A7:D7')
        
        # Column headers
        headers = ['Category', 'Item Count', 'Est. Hours', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Placeholder data
        categories = [
            ['Equipment', '=Equipment!B100', '=Equipment!C100', 'See Equipment tab'],
            ['Wire & Cable', '=\'Wire & Cable\'!B100', '=\'Wire & Cable\'!C100', 'See Wire tab'],
            ['Conduit', '=\'Conduit & Fittings\'!B100', '=\'Conduit & Fittings\'!C100', 'See Conduit tab']
        ]
        
        for row, cat_data in enumerate(categories, 9):
            for col, value in enumerate(cat_data, 1):
                ws.cell(row=row, column=col).value = value
        
        # Format columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
    
    def _create_equipment_sheet(self, ws, equipment_list):
        """Create equipment takeoff sheet"""
        # Header
        ws['A1'] = "ELECTRICAL EQUIPMENT"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:G1')
        
        # Column headers
        headers = ['Item', 'Description', 'Rating', 'Qty', 'Unit', 'Labor Hrs', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border
        
        # Add equipment items
        row = 3
        for item in equipment_list:
            ws.cell(row=row, column=1).value = item.get('item_number', '')
            ws.cell(row=row, column=2).value = item.get('description', '')
            ws.cell(row=row, column=3).value = item.get('rating', '')
            ws.cell(row=row, column=4).value = item.get('quantity', 1)
            ws.cell(row=row, column=5).value = item.get('unit', 'EA')
            ws.cell(row=row, column=6).value = item.get('labor_hours', 0)
            ws.cell(row=row, column=7).value = item.get('notes', '')
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.thin_border
            
            row += 1
        
        # Totals
        ws.cell(row=100, column=2).value = "TOTAL ITEMS:"
        ws.cell(row=100, column=2).font = self.total_font
        ws.cell(row=100, column=4).value = f"=SUM(D3:D99)"
        ws.cell(row=100, column=4).font = self.total_font
        ws.cell(row=100, column=6).value = f"=SUM(F3:F99)"
        ws.cell(row=100, column=6).font = self.total_font
        
        # Format columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
    
    def _create_wire_sheet(self, ws, wire_list):
        """Create wire & cable takeoff sheet"""
        # Header
        ws['A1'] = "WIRE & CABLE"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:G1')
        
        # Column headers
        headers = ['Item', 'Size', 'Type', 'Qty', 'Unit', 'Labor Hrs', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border
        
        # Add wire items
        row = 3
        for item in wire_list:
            ws.cell(row=row, column=1).value = item.get('item_number', '')
            ws.cell(row=row, column=2).value = item.get('size', '')
            ws.cell(row=row, column=3).value = item.get('type', 'THHN/THWN')
            ws.cell(row=row, column=4).value = item.get('quantity', 0)
            ws.cell(row=row, column=5).value = item.get('unit', 'FT')
            ws.cell(row=row, column=6).value = item.get('labor_hours', 0)
            ws.cell(row=row, column=7).value = item.get('notes', '')
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.thin_border
            
            row += 1
        
        # Totals
        ws.cell(row=100, column=2).value = "TOTAL FOOTAGE:"
        ws.cell(row=100, column=2).font = self.total_font
        ws.cell(row=100, column=4).value = f"=SUM(D3:D99)"
        ws.cell(row=100, column=4).font = self.total_font
        ws.cell(row=100, column=6).value = f"=SUM(F3:F99)"
        ws.cell(row=100, column=6).font = self.total_font
        
        # Format columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
    
    def _create_conduit_sheet(self, ws, conduit_list):
        """Create conduit & fittings takeoff sheet"""
        # Similar structure to wire sheet
        ws['A1'] = "CONDUIT & FITTINGS"
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        ws.merge_cells('A1:G1')
        
        headers = ['Item', 'Size', 'Type', 'Qty', 'Unit', 'Labor Hrs', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border
        
        row = 3
        for item in conduit_list:
            ws.cell(row=row, column=1).value = item.get('item_number', '')
            ws.cell(row=row, column=2).value = item.get('size', '')
            ws.cell(row=row, column=3).value = item.get('type', 'EMT')
            ws.cell(row=row, column=4).value = item.get('quantity', 0)
            ws.cell(row=row, column=5).value = item.get('unit', 'FT')
            ws.cell(row=row, column=6).value = item.get('labor_hours', 0)
            ws.cell(row=row, column=7).value = item.get('notes', '')
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self.thin_border
            
            row += 1
        
        ws.cell(row=100, column=2).value = "TOTAL:"
        ws.cell(row=100, column=2).font = self.total_font
        ws.cell(row=100, column=4).value = f"=SUM(D3:D99)"
        ws.cell(row=100, column=4).font = self.total_font
        
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
    
    def _parse_electrical_analysis(self, analysis: str) -> tuple:
        """
        Parse Claude's analysis to extract material quantities
        
        Returns: (equipment_list, wire_list, conduit_list)
        """
        equipment = []
        wire = []
        conduit = []
        
        # Extract main switchboard
        if '800A' in analysis or '800AMP' in analysis:
            equipment.append({
                'item_number': 'SB-01',
                'description': 'Main Switchboard',
                'rating': '800A, 480V',
                'quantity': 1,
                'unit': 'EA',
                'labor_hours': 40,
                'notes': 'Main service equipment'
            })
        
        # Extract buckets/panels
        bucket_pattern = r'(\d+)AT.*?(\d+)A'
        for match in re.finditer(bucket_pattern, analysis):
            trip_rating = match.group(1)
            equipment.append({
                'item_number': f'PNL-{len(equipment)}',
                'description': f'Panel, {trip_rating}A',
                'rating': f'{trip_rating}A',
                'quantity': 1,
                'unit': 'EA',
                'labor_hours': 8,
                'notes': f'Distribution panel'
            })
        
        # Extract wire
        if '600 kcmil' in analysis.lower() or '600kcmil' in analysis.lower():
            wire.append({
                'item_number': 'W-01',
                'size': '600 kcmil',
                'type': 'THHN/THWN',
                'quantity': 500,  # Estimated
                'unit': 'FT',
                'labor_hours': 20,
                'notes': 'Main feeder - parallel sets'
            })
        
        if '#1/0' in analysis or '1/0' in analysis:
            wire.append({
                'item_number': 'W-02',
                'size': '#1/0 AWG',
                'type': 'THHN/THWN',
                'quantity': 300,  # Estimated
                'unit': 'FT',
                'labor_hours': 12,
                'notes': 'Ground wire'
            })
        
        return equipment, wire, conduit


def export_to_excel(analysis_result: str, drawing_name: str, project_name: str = "Construction Project") -> str:
    """
    Convenience function to export analysis to Excel
    
    Args:
        analysis_result: Claude's analysis text
        drawing_name: Name of the drawing
        project_name: Project name for the takeoff
    
    Returns:
        Path to Excel file
    """
    exporter = MaterialTakeoffExporter(project_name)
    return exporter.create_electrical_takeoff(analysis_result, drawing_name)
